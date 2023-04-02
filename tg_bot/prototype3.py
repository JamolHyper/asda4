import logging
from datetime import datetime
from telegram import Update
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, ConversationHandler

# Set up logging
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)

# Create an Excel workbook to store the reports
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active
ws.title = "Fuel Reports"

# Add headers to the worksheet
headers = ['ФИО', 'Сана', 'Трактор', 'Смена', 'Оператор', 'Иш бошлангандаги мотор соати', 
           'Ёкилги микдори', 'Ёкилги олинган вакти', 'Иш тугагандан кейинги мотор соати']
for i, header in enumerate(headers):
    cell = ws.cell(row=1, column=i+1)
    cell.value = header
    cell.font = cell.font.copy(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Define the conversation flow
def start(update: Update, context):
    """Start the conversation and ask for the user's full name."""
    update.message.reply_text("Хуш келибсиз! Исм шарифингизни киритинг.")
    return 'name'

def get_name(update: Update, context):
    """Store the user's full name and ask for the date."""
    context.user_data['name'] = update.message.text
    update.message.reply_text("Санани киритинг.")
    return 'date'

def get_date(update: Update, context):
    """Store the date and ask for the tractor model."""
    try:
        date = datetime.strptime(update.message.text, '%d.%m.%Y' or '%d %m %Y').date()
        context.user_data['date'] = date
        update.message.reply_text("Тракторни танланг:",
                                  reply_markup=get_keyboard(['LD-8001', 'LD-8002', 'LD-8003', 'LD-8004', 'LD-8005',
                                                             'LD-8006', 'LD-8007', 'LD-8008', 'LD-8009', 'LD-8010',
                                                             'LD-8011', 'LD-8012', 'LD-8013', 'LD-8014', 'LD-8015',
                                                             'LD-9001', 'LD-9002', 'LD-9003', 'LD-9004', 'LD-9005',
                                                             'LD-9006', 'LD-9007', 'LD-9008', 'LD-9009', 'LD-9010',
                                                             'LD-9011', 'LD-9012']))
        return 'tractor'
    except ValueError:
        update.message.reply_text("кун.ой.йил форматида киритинг (мисол учун: 18.03.2023)")
        return 'date'

def get_tractor(update: Update, context):
    """Store the tractor model and ask for the shift."""
    context.user_data['tractor'] = update.message.text
    update.message.reply_text("Сменани танланг:", reply_markup=get_keyboard(['Кундузги', 'Кечги']))
    return 'shift'

def get_shift(update: Update, context):
    """Store the shift and ask for the amount of fuel."""
    context.user_data['shift'] = update.message.text
    update.message.reply_text("Операторни танланг.", reply_markup=get_keyboard(['Оператор 1', 'Оператор 2', 'Оператор 3']))
    return 'operator'

def get_operator(update: Update, context):
    context.user_data['operator'] = update.message.text
    update.message.reply_text("Иш бошлангандаги мотор соатини киритинг.")
    return 'motor_hours_before'

def get_motor_hours_before(update: Update, context):
    """Store the amount of fuel and ask for the time it was filled."""
    try:
        motor_hours_before = int(update.message.text)
        context.user_data['motor_hours_before'] = motor_hours_before
        update.message.reply_text("Ёкилги микдорини киритинг")
        return 'fuel'
    except ValueError:
        update.message.reply_text("Хатолик. Мотор соатини кайта киритинг.")
    return 'motor_hours_before'

def get_fuel(update: Update, context):
    """Store the amount of fuel and ask for the time it was filled."""
    try:
        fuel = int(update.message.text)
        context.user_data['fuel'] = fuel
        update.message.reply_text("Качон ёкилгини олганингизни киритинг.")
        return 'time'
    except ValueError:
        update.message.reply_text("Илтимос, ёкилги микдорини тугри киритинг.")
    return 'fuel'

def get_time(update: Update, context):
    """Store the time the fuel was filled and ask for the motor hours at the end of the day."""
    try:
        time = datetime.strptime(update.message.text, '%H:%M').time()
        context.user_data['time'] = time
        update.message.reply_text("Иш тугагандаги мотор соатини киритинг.")
        return 'motor_hours'
    except ValueError:
        update.message.reply_text("Илтимос вактни тугри киритинг (мисол учун: 18:30).")
    return 'time'

def get_motor_hours(update: Update, context):
    """Store the motor hours and save the report to the Excel worksheet."""
    try:
        motor_hours = float(update.message.text)
        context.user_data['motor_hours'] = motor_hours
# Add the report to the worksheet
        row = [context.user_data['name'], context.user_data['date'].strftime('%d.%m.%Y'),
        context.user_data['tractor'], context.user_data['shift'], context.user_data['operator'], context.user_data['motor_hours_before'], context.user_data['fuel'],
        context.user_data['time'].strftime('%H:%M' or '%H %M'), context.user_data['motor_hours'], ]
        ws.append(row)
        wb.save('fuel_reports.xlsx')
# Send a confirmation message
        update.message.reply_text("Хисоботингиз кабул килинди!")
# Reset the conversation
        return start(update, context)
    except ValueError:
        update.message.reply_text("Мотор соатини бутун сонда киритинг.")
    return 'motor_hours'

def cancel(update: Update, context):
    """Cancel the conversation and reset the user data."""
    update.message.reply_text("Бекор килинди.")
    return ConversationHandler.END

def get_keyboard(options):
    """Create a keyboard with the given options."""
    keyboard = [[option] for option in options]
    return {'keyboard': keyboard, 'one_time_keyboard': True}

def main():
    """Start the bot."""
# Create an Updater instance and connect to the Telegram API
updater = Updater("1961254009:AAGvd138v3bYRRbOAr_X_g73o3YgsIoR-ZI", use_context=True)

# Get the dispatcher to register handlers
dispatcher = updater.dispatcher

# Set up a conversation handler
conv_handler = ConversationHandler(
    entry_points=[CommandHandler('start', start)],
    states={
        'name': [MessageHandler(Filters.text, get_name)],
        'date': [MessageHandler(Filters.text, get_date)],
        'tractor': [MessageHandler(Filters.text, get_tractor)],
        'shift': [MessageHandler(Filters.text, get_shift)],
        'operator': [MessageHandler(Filters.text, get_operator)],
        'fuel': [MessageHandler(Filters.text, get_fuel)],
        'time': [MessageHandler(Filters.text, get_time)],
        'motor_hours': [MessageHandler(Filters.text, get_motor_hours)],
        'motor_hours_before':[MessageHandler(Filters.text, get_motor_hours_before)],
    },
    fallbacks=[CommandHandler('cancel', cancel)]
)

# Add the conversation handler to the dispatcher
dispatcher.add_handler(conv_handler)

# Start the bot
updater.start_polling()

# Run the bot until the user presses Ctrl-C or the process
